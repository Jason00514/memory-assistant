"""
Parse Excel files into RawContent records.

Column format (single column):
  # Category: <name>   -> sets category for subsequent cells
  <cell content>       -> one record per non-empty cell
"""
import re
from typing import BinaryIO
from openpyxl import load_workbook


def detect_content_type(text: str) -> str:
    if "answer check:" in text:
        return "multiple_choice"
    if "answer option:" in text:
        return "single_choice"
    if "answer:" in text:
        return "word"
    return "memory"


def parse_excel_stream(file: BinaryIO) -> list[dict]:
    """
    Read the first sheet of an Excel file and return a list of raw record dicts.
    Each dict has: category, content_type, raw_text
    """
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active

    records = []
    current_category = "未分类"
    category_pattern = re.compile(r"#\s*[Cc]ategory\s*:\s*(.+)", re.IGNORECASE)

    for row in ws.iter_rows(values_only=True):
        cell_value = row[0] if row else None
        if cell_value is None:
            continue
        text = str(cell_value).strip()
        if not text:
            continue

        m = category_pattern.match(text)
        if m:
            current_category = m.group(1).strip()
            continue

        content_type = detect_content_type(text)
        records.append({
            "category": current_category,
            "content_type": content_type,
            "raw_text": text,
        })

    wb.close()
    return records
